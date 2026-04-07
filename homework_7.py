import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from typing import cast
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# Создаем новый файл для домашки
wb = openpyxl.Workbook()

# Настраиваем стиль тонких границ для ячеек
thin_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# ==========================================
# ДОМАШНЕЕ ЗАДАНИЕ 7 - ЗАДАНИЕ 1
# ==========================================
ws1 = cast(Worksheet, wb.active)
ws1.title = "ДЗ 7 - Задание 1"

# 1. Шапка таблицы
headers_task1 = ['Город', 'Население города', 'Население страны', 'Доля города от страны, %']
header_fill_t1 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

for col_idx, text in enumerate(headers_task1, start=1):
    cell = cast(Cell, ws1.cell(row=1, column=col_idx))
    cell.value = text
    cell.fill = header_fill_t1
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 2. Данные 10 крупнейших городов Европы (население примерное, для демонстрации)
cities_data = [
    ['Стамбул', 15460000, 85200000],
    ['Москва', 13010000, 144200000],
    ['Лондон', 8980000, 67300000],
    ['Санкт-Петербург', 5600000, 144200000],
    ['Берлин', 3769000, 83200000],
    ['Мадрид', 3223000, 47400000],
    ['Киев', 2962000, 38000000],
    ['Рим', 2872000, 59000000],
    ['Париж', 2161000, 67900000],
    ['Бухарест', 1836000, 19000000]
]

# 3. Заполняем данные и прописываем границы
for i, row_data in enumerate(cities_data, start=2):
    cast(Cell, ws1.cell(row=i, column=1)).value = row_data[0] # Город
    cast(Cell, ws1.cell(row=i, column=2)).value = row_data[1] # Население города
    cast(Cell, ws1.cell(row=i, column=3)).value = row_data[2] # Население страны
    
    # Формула: Население города (B) / Население страны (C)
    cast(Cell, ws1.cell(row=i, column=4)).value = f'=B{i}/C{i}'

# 4. Настраиваем форматы и применяем границы ко всей таблице
for row in ws1['A1:D11']:
    for cell in row:
        c = cast(Cell, cell)
        c.border = thin_border # Рисуем границы!
        
        # Настраиваем формат чисел (разделитель тысяч для людей, проценты для доли)
        if c.column in [2, 3] and c.row > 1:
            c.number_format = '#,##0'
        elif c.column == 4 and c.row > 1:
            c.number_format = '0.0%'

ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 20


# ==========================================
# ДОМАШНЕЕ ЗАДАНИЕ 7 - ЗАДАНИЕ 2 (Аптека)
# ==========================================
ws2 = cast(Worksheet, wb.create_sheet(title="ДЗ 7 - Задание 2"))

# 1. Сложная шапка с объединением ячеек
dark_blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)

headers_task2 = [
    ('A1', '№', 'A1:A2'),
    ('B1', 'Наименование', 'B1:B2'),
    ('C1', 'Закупочная\nцена', 'C1:C2'),
    ('D1', 'Наценка', 'D1:E1'), # Наценка объединяет D и E по горизонтали
    ('F1', 'Цена\nрозничная', 'F1:F2'),
    ('G1', 'Количество', 'G1:G2'),
    ('H1', 'Сумма', 'H1:H2')
]

for cell_ref, text, merge_range in headers_task2:
    cell = cast(Cell, ws2[cell_ref])
    cell.value = text
    ws2.merge_cells(merge_range)

# Подзаголовки для Наценки
cast(Cell, ws2['D2']).value = '%'
cast(Cell, ws2['E2']).value = 'руб.'

# Красим всю шапку в темно-синий и делаем текст белым
for row in ws2['A1:H2']:
    for cell in row:
        c = cast(Cell, cell)
        c.fill = dark_blue_fill
        c.font = white_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border

# 2. Вносим данные (проценты снова пишем как дроби: 45% = 0.45)
pharmacy_data = [
    [1, 'Амлодипин таб. 5 мг №30', 11.85, 0.45, 15],
    [2, 'Амлодипин таб. 10 мг №30', 14.25, 0.45, 40],
    [3, 'Адвантан мазь д/нар. прим. 15 г', 337.76, 0.38, 11],
    [4, 'Окопник растирка 150 мл', 76.85, 0.40, 5]
]

for i, row_data in enumerate(pharmacy_data, start=3):
    cast(Cell, ws2.cell(row=i, column=1)).value = row_data[0] # №
    cast(Cell, ws2.cell(row=i, column=2)).value = row_data[1] # Наименование
    cast(Cell, ws2.cell(row=i, column=3)).value = row_data[2] # Закупочная цена
    cast(Cell, ws2.cell(row=i, column=4)).value = row_data[3] # Наценка %
    
    # Формула: Наценка руб. (E) = Наценка % (D) * Закупочная цена (C)
    cast(Cell, ws2.cell(row=i, column=5)).value = f'=D{i}*C{i}'
    
    # Формула: Цена розничная (F) = Закупочная цена (C) + Наценка руб. (E)
    cast(Cell, ws2.cell(row=i, column=6)).value = f'=C{i}+E{i}'
    
    cast(Cell, ws2.cell(row=i, column=7)).value = row_data[4] # Количество
    
    # Формула: Сумма (H) = Цена розничная (F) * Количество (G)
    cast(Cell, ws2.cell(row=i, column=8)).value = f'=F{i}*G{i}'

# 3. Строка "ИТОГО"
last_row = 7
ws2.merge_cells(f'A{last_row}:G{last_row}')
cell_total = cast(Cell, ws2[f'A{last_row}'])
cell_total.value = 'ИТОГО:'
cell_total.alignment = Alignment(horizontal='right', vertical='center')

# Формула автосуммы для колонки "Сумма" (H)
cast(Cell, ws2[f'H{last_row}']).value = f'=SUM(H3:H{last_row-1})'

# Красим строку ИТОГО в светло-голубой
light_blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
for col_idx in range(1, 9):
    cast(Cell, ws2.cell(row=last_row, column=col_idx)).fill = light_blue_fill

# 4. Форматы и границы для данных
for row in ws2[f'A3:H{last_row}']:
    for cell in row:
        c = cast(Cell, cell)
        c.border = thin_border
        
        # Денежный формат для столбцов C, E, F, H (строки 3-6)
        if c.column in [3, 5, 6, 8] and c.row < last_row:
            c.number_format = '#,##0.00 "₽"'
        # Денежный формат для итоговой суммы (строка 7)
        if c.column == 8 and c.row == last_row:
            c.number_format = '#,##0.00 "₽"'
        # Процентный формат для столбца D (строки 3-6)
        if c.column == 4 and c.row < last_row:
            c.number_format = '0%'

# 5. Ширина столбцов
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 15

# ==========================================
# СОХРАНЕНИЕ
# ==========================================
filename = 'Домашнее задание 7.xlsx'
wb.save(filename)
print(f'Бро, всё готово! Файл "{filename}" успешно создан.')
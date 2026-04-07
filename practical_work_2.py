# Импортируем библиотеку для работы с Excel
import openpyxl
# Импортируем инструменты для настройки внешнего вида ячеек
from openpyxl.styles import Alignment, Font, PatternFill

# Импортируем типы данных, чтобы подсказать Pyright, с чем мы работаем
from typing import cast
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.formatting.rule import CellIsRule

# 1. Создаем новую рабочую книгу (файл)
wb = openpyxl.Workbook()

# Успокаиваем Pyright: жестко указываем, что активный лист — это объект Worksheet
ws = cast(Worksheet, wb.active)
ws.title = "Практическая работа 2"

# 2. Оформляем главный заголовок таблицы
# Pyright сомневается, одна это ячейка или диапазон, поэтому применяем cast(Cell, ...)
cell_a1 = cast(Cell, ws['A1'])
cell_a1.value = 'Анализ спроса и продаж продукции фирмы "Шанс"'

ws.merge_cells('A1:H1') # Объединяем ячейки от A до H
cell_a1.font = Font(bold=True, size=12)
cell_a1.alignment = Alignment(horizontal='center', vertical='center')

# 3. Настраиваем шапку таблицы (названия столбцов)
headers = [
    ('A2', 'Наименование\nпродукции', 'A2:A3'),
    ('B2', 'Цена за\nед.', 'B2:B3'),
    ('C2', 'Спрос,\nшт.', 'C2:C3'),
    ('D2', 'Предл.,\nшт.', 'D2:D3'),
    ('E2', 'Продажа', 'E2:G2'), 
    ('H2', 'Выручка\nот\nпродаж', 'H2:H3')
]

for cell_ref, text, merge_range in headers:
    current_cell = cast(Cell, ws[cell_ref])
    current_cell.value = text
    ws.merge_cells(merge_range)

# Добавляем подзаголовки
cast(Cell, ws['E3']).value = 'Безнал.'
cast(Cell, ws['F3']).value = 'Нал.'
cast(Cell, ws['G3']).value = 'Всего'

# Применяем форматирование ко всем ячейкам шапки
header_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")

for row in ws['A2:H3']:
    for cell in row:
        # В циклах по строкам получаем ячейки, тоже подсказываем их тип
        c = cast(Cell, cell)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.font = Font(bold=True)
        c.fill = header_fill

# 4. Вносим исходные данные
data = [
    ['Телевизоры', 350.35, 13, 15, 5, 7],
    ['Видеомагнитофоны', 320.00, 70, 65, 30, 35],
    ['Проигрыватели', 400.21, 65, 134, 40, 26]
]

# Метод .cell() более безопасен для типов, так что тут Pyright ругаться не должен
for i, row_data in enumerate(data, start=4):
    ws.cell(row=i, column=1, value=row_data[0]) 
    ws.cell(row=i, column=2, value=row_data[1]) 
    ws.cell(row=i, column=3, value=row_data[2]) 
    ws.cell(row=i, column=4, value=row_data[3]) 
    ws.cell(row=i, column=5, value=row_data[4]) 
    ws.cell(row=i, column=6, value=row_data[5]) 
    
    # 5. Прописываем формулы Excel
    ws.cell(row=i, column=7, value=f'=E{i}+F{i}')
    ws.cell(row=i, column=8, value=f'=B{i}*F{i}')

# 6. Применяем денежный формат
for row_num in range(4, 7):
    cast(Cell, ws[f'B{row_num}']).number_format = '$ #,##0.00'
    cast(Cell, ws[f'H{row_num}']).number_format = '$ #,##0.00'

# Настраиваем ширину столбцов
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 15

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 2 ---

# 1. Создаем новый лист в нашей книге
# Опять же, используем cast, чтобы Pyright понимал, что это Worksheet
ws2 = cast(Worksheet, wb.create_sheet(title="Задание 2"))

# 2. Оформляем главный заголовок таблицы
cell_a1_ws2 = cast(Cell, ws2['A1'])
cell_a1_ws2.value = 'РАСХОД МАТЕРИАЛОВ ДЛЯ ОКРАСКИ'
ws2.merge_cells('A1:G1') # Объединяем на 7 колонок (от A до G)
cell_a1_ws2.font = Font(bold=True, size=12)
cell_a1_ws2.alignment = Alignment(horizontal='center', vertical='center')

# 3. Настраиваем сложную многоуровневую шапку
headers_ws2 = [
    ('A2', 'Материал', 'A2:A4'),       # 3 ячейки по вертикали
    ('B2', 'Поверхность', 'B2:G2'),    # 6 ячеек по горизонтали
    ('B3', 'Двери', 'B3:D3'),          # 3 ячейки по горизонтали
    ('E3', 'Окна', 'E3:G3')            # 3 ячейки по горизонтали
]

for cell_ref, text, merge_range in headers_ws2:
    current_cell = cast(Cell, ws2[cell_ref])
    current_cell.value = text
    ws2.merge_cells(merge_range)

# Добавляем подзаголовки 4-й строки. 
# Заметь: вместо сложных шрифтов мы используем \u00B2 — это символ квадрата '²'
subheaders = ['Кг на\n10 м\u00B2', 'Площадь', 'Расход', 'Кг на\n10 м\u00B2', 'Площадь', 'Расход']
for col_idx, text in enumerate(subheaders, start=2): # Начинаем со 2-го столбца (B)
    cell = cast(Cell, ws2.cell(row=4, column=col_idx))
    cell.value = text

# Применяем выравнивание по центру ко всей шапке Задания 2
for row in ws2['A2:G4']:
    for cell in row:
        c = cast(Cell, cell)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.font = Font(bold=True)

# 4. Вносим исходные данные
# Важно: в Python дробные числа пишутся через точку, а не через запятую
data_ws2 = [
    ['Олифа', 1.20, 150, 1.35, 362],
    ['Белила', 0.87, 150, 0.91, 362],
    ['Пигмент', 0.15, 150, 0.42, 362]
]

# Записываем данные и формулы начиная с 5-й строки
for i, row_data in enumerate(data_ws2, start=5):
    cast(Cell, ws2.cell(row=i, column=1)).value = row_data[0] # Материал
    
    # --- Заполняем секцию "Двери" ---
    cast(Cell, ws2.cell(row=i, column=2)).value = row_data[1] # Кг на 10 м2
    cast(Cell, ws2.cell(row=i, column=3)).value = row_data[2] # Площадь
    
    # Формула: Расход (Двери) = Кг (B) * Площадь (C) / 10
    cast(Cell, ws2.cell(row=i, column=4)).value = f'=B{i}*C{i}/10'
    
    # --- Заполняем секцию "Окна" ---
    cast(Cell, ws2.cell(row=i, column=5)).value = row_data[3] # Кг на 10 м2
    cast(Cell, ws2.cell(row=i, column=6)).value = row_data[4] # Площадь
    
    # Формула: Расход (Окна) = Кг (E) * Площадь (F) / 10
    cast(Cell, ws2.cell(row=i, column=7)).value = f'=E{i}*F{i}/10'

# 5. Делаем столбцы чуть шире, чтобы текст не слипался
ws2.column_dimensions['A'].width = 15
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws2.column_dimensions[col].width = 11

# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 2 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 3 ---

# 1. Создаем третий лист
ws3 = cast(Worksheet, wb.create_sheet(title="Задание 3"))

# 2. Оформляем заголовок "Валовая прибыль"
cell_a1_ws3 = cast(Cell, ws3['A1'])
cell_a1_ws3.value = 'Валовая прибыль'
ws3.merge_cells('A1:H1')
cell_a1_ws3.alignment = Alignment(horizontal='center', vertical='center')
cell_a1_ws3.font = Font(bold=True, size=12)

# 3. Настраиваем шапку и боковик (названия строк)
months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Всего']
for col_idx, month in enumerate(months, start=2): # Столбцы с B(2) по H(8)
    cast(Cell, ws3.cell(row=2, column=col_idx)).value = month

row_labels = ['Доходы', 'Расходы', 'Прибыль']
for row_idx, label in enumerate(row_labels, start=3): # Строки с 3 по 5
    cast(Cell, ws3.cell(row=row_idx, column=1)).value = label

# 4. Красим шапку и боковик в светло-зеленый цвет (как на картинке)
green_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

# Красим верхнюю шапку
for row in ws3['A2:H2']:
    for cell in row:
        c = cast(Cell, cell)
        c.fill = green_fill
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

# Красим левый боковик
for row in ws3['A3:A5']:
    for cell in row:
        c = cast(Cell, cell)
        c.fill = green_fill
        c.font = Font(bold=True)

# 5. Вносим исходные данные
incomes = [112100, 112850, 148800, 106500, 101150, 79700]
expenses = [90567, 95100, 105900, 129200, 131000, 84000]

cols = ['B', 'C', 'D', 'E', 'F', 'G']

for i, col in enumerate(cols):
    cast(Cell, ws3[f'{col}3']).value = incomes[i]
    cast(Cell, ws3[f'{col}4']).value = expenses[i]
    
    # Считаем прибыль: "Доходы" - "Расходы"
    cast(Cell, ws3[f'{col}5']).value = f'={col}3-{col}4'

# 6. Считаем столбец "Всего" (Сумма по строке)
for row_idx in [3, 4, 5]:
    # Формула суммы от колонки B до G для текущей строки
    cast(Cell, ws3[f'H{row_idx}']).value = f'=SUM(B{row_idx}:G{row_idx})'

# 7. Применяем формат "Рубли" ко всем числам
for row in ws3['B3:H5']:
    for cell in row:
        # Формат с пробелом-разделителем тысяч и знаком рубля
        cast(Cell, cell).number_format = '#,##0 "₽"'

# 8. УСЛОВНОЕ ФОРМАТИРОВАНИЕ (Магия из задания)
# Создаем правило: если значение < 0, то красим в светло-красный (и делаем текст темно-красным для красоты)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006")

rule = CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font)

# Применяем правило к строке с прибылью (от января до 'Всего')
ws3.conditional_formatting.add('B5:H5', rule)

# Настраиваем ширину столбцов
ws3.column_dimensions['A'].width = 12
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws3.column_dimensions[col].width = 12

# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 3 ---


filename = 'Практическая работа 2.xlsx'
wb.save(filename)
print(f'Готово! Файл "{filename}" успешно обновлен.')

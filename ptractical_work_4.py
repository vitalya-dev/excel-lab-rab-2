import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, PieChart3D, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from typing import cast
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# Создаем рабочую книгу
wb = openpyxl.Workbook()

# ==========================================
# ЗАДАНИЕ 1: ГИСТОГРАММА ПЛАН/ФАКТ
# ==========================================
ws1 = cast(Worksheet, wb.active)
ws1.title = "Задание 1"

cell_a1 = cast(Cell, ws1['A1'])
cell_a1.value = 'Показатели производства за 2014 год'
ws1.merge_cells('A1:E1')
cell_a1.font = Font(bold=True)
cell_a1.alignment = Alignment(horizontal='center', vertical='center')

headers_ws1 = ['1', '2', '3', '4']
for col_idx, text in enumerate(headers_ws1, start=2):
    c = cast(Cell, ws1.cell(row=2, column=col_idx))
    c.value = text
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')

data_ws1 = [
    ['План (тыс. руб)', 1000, 800, 1500, 1100],
    ['Факт (тыс. руб)', 980, 1150, 1200, 1060]
]

for i, row_data in enumerate(data_ws1, start=3):
    for j, val in enumerate(row_data, start=1):
        cell = cast(Cell, ws1.cell(row=i, column=j))
        cell.value = val
        if j == 1: cell.font = Font(bold=True)
        else: cell.alignment = Alignment(horizontal='center')

ws1.column_dimensions['A'].width = 18

chart1 = BarChart()
chart1.type = "col"
chart1.title = "Поквартальное выполнение плана"
data_ref1 = Reference(ws1, min_col=1, min_row=3, max_col=5, max_row=4)
cats_ref1 = Reference(ws1, min_col=2, min_row=2, max_col=5, max_row=2)
chart1.add_data(data_ref1, titles_from_data=True, from_rows=True)
chart1.set_categories(cats_ref1)
ws1.add_chart(chart1, "A6")

# ==========================================
# ЗАДАНИЕ 2: ОБЪЕМНАЯ КРУГОВАЯ ДИАГРАММА
# ==========================================
ws2 = cast(Worksheet, wb.create_sheet(title="Задание 2"))

headers_ws2 = ['1 квартал', '2 квартал', '3 квартал', '4 квартал']
for col_idx, text in enumerate(headers_ws2, start=2):
    c = cast(Cell, ws2.cell(row=1, column=col_idx))
    c.value = text
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')

data_ws2 = ['Факт (тыс.руб.)', 980, 1150, 1200, 1060]
for col_idx, val in enumerate(data_ws2, start=1):
    c = cast(Cell, ws2.cell(row=2, column=col_idx))
    c.value = val
    if col_idx == 1: c.font = Font(bold=True)

ws2.column_dimensions['A'].width = 18
for col in ['B', 'C', 'D', 'E']:
    ws2.column_dimensions[col].width = 12

pie = PieChart3D()
pie.title = "Фактическое выполнение плана"
data_ref2 = Reference(ws2, min_col=2, min_row=2, max_col=5, max_row=2)
cats_ref2 = Reference(ws2, min_col=2, min_row=1, max_col=5, max_row=1)
pie.add_data(data_ref2, from_rows=True)
pie.set_categories(cats_ref2)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws2.add_chart(pie, "A4")

# ==========================================
# ЗАДАНИЯ 3, 4, 5, 6: МАТЕМАТИЧЕСКИЕ ФОРМУЛЫ
# ==========================================
ws3 = cast(Worksheet, wb.create_sheet(title="Формулы (4-6)"))

ws3['A1'], ws3['B1'] = 'X', 'Y'
ws3['A2'], ws3['B2'] = 4, 3  

tasks = [('C1', 'Задание 3'), ('C2', 'Задание 4'), ('C3', 'Задание 5'), ('C4', 'Задание 6')]

for cell_ref, label in tasks:
    c = cast(Cell, ws3[cell_ref])
    c.value = label
    c.font = Font(bold=True)

ws3['D1'] = '=(1+A2)/(4*B2)'
ws3['D2'] = '=-2*A2+(A2^5)/(3*B2^2+4)'
ws3['D3'] = '=SQRT(7*A2+2)'
ws3['D4'] = '=SIN((A2+5)/(3*A2-2))+SQRT(A2^3+1)'

for row_idx in range(1, 5):
    res_cell = cast(Cell, ws3.cell(row=row_idx, column=4))
    res_cell.number_format = '0.000'

for cell_ref in ['A1', 'B1']:
    c = cast(Cell, ws3[cell_ref])
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

ws3.column_dimensions['C'].width, ws3.column_dimensions['D'].width = 12, 12

# ==========================================
# ЗАДАНИЕ 7: y = sin(2.5(x-3))
# ==========================================
ws7 = cast(Worksheet, wb.create_sheet(title="Задание 7"))
ws7['A1'], ws7['B1'] = 'X', 'Y'
ws7['A1'].font = ws7['B1'].font = Font(bold=True)

current_x, current_row = -4.0, 2
while current_x <= 4.01:
    ws7.cell(row=current_row, column=1, value=current_x).number_format = '0.00'
    ws7.cell(row=current_row, column=2, value=f'=SIN(2.5*(A{current_row}-3))').number_format = '0.00'
    current_x = round(current_x + 0.2, 2)
    current_row += 1

chart7 = ScatterChart()
chart7.title = "y=sin 2.5(x-3)"
chart7.legend = None  # ОТКЛЮЧАЕМ ЛЕГЕНДУ

series7 = Series(Reference(ws7, min_col=2, min_row=2, max_row=current_row-1), 
                 Reference(ws7, min_col=1, min_row=2, max_row=current_row-1))
series7.smooth = True 
chart7.series.append(series7)
ws7.add_chart(chart7, "D2")

# ==========================================
# ЗАДАНИЕ 8: y = log3(x+1) - ФИОЛЕТОВЫЙ ПУНКТИР
# ==========================================
ws8 = cast(Worksheet, wb.create_sheet(title="Задание 8"))
ws8['A1'], ws8['B1'] = 'X', 'Y'
ws8['A1'].font = ws8['B1'].font = Font(bold=True)

current_x, current_row = -0.8, 2
while current_x <= 3.01:
    ws8.cell(row=current_row, column=1, value=current_x).number_format = '0.00'
    ws8.cell(row=current_row, column=2, value=f'=LOG(A{current_row}+1, 3)').number_format = '0.00'
    current_x = round(current_x + 0.2, 2)
    current_row += 1

chart8 = ScatterChart()
chart8.title = "y=log3(x+1)"
chart8.legend = None  # ОТКЛЮЧАЕМ ЛЕГЕНДУ

chart8.x_axis.scaling.min, chart8.x_axis.scaling.max = -2.0, 4.0
chart8.y_axis.scaling.min, chart8.y_axis.scaling.max = -3.0, 2.0

series8 = Series(Reference(ws8, min_col=2, min_row=2, max_row=current_row-1), 
                 Reference(ws8, min_col=1, min_row=2, max_row=current_row-1))
series8.smooth = True 
series8.graphicalProperties.line.solidFill = "7030A0" # Фиолетовый
series8.graphicalProperties.line.dashStyle = "dash"   # Пунктир
series8.graphicalProperties.line.width = 30000 

chart8.series.append(series8)
ws8.add_chart(chart8, "D2")

# ==========================================
# ЗАДАНИЕ 9: y = x * cos(x) - ЗЕЛЕНЫЙ
# ==========================================
ws9 = cast(Worksheet, wb.create_sheet(title="Задание 9"))
ws9['A1'], ws9['B1'] = 'X', 'Y'
ws9['A1'].font = ws9['B1'].font = Font(bold=True)

current_x, current_row = -10.0, 2
while current_x <= 10.01:
    ws9.cell(row=current_row, column=1, value=current_x).number_format = '0.00'
    ws9.cell(row=current_row, column=2, value=f'=A{current_row}*COS(A{current_row})').number_format = '0.00'
    current_x = round(current_x + 0.1, 2)
    current_row += 1

chart9 = ScatterChart()
chart9.title = "y = x * cos(x)"
chart9.legend = None  # ОТКЛЮЧАЕМ ЛЕГЕНДУ

chart9.x_axis.scaling.min, chart9.x_axis.scaling.max = -12.0, 12.0
chart9.y_axis.scaling.min, chart9.y_axis.scaling.max = -12.0, 12.0

series9 = Series(Reference(ws9, min_col=2, min_row=2, max_row=current_row-1), 
                 Reference(ws9, min_col=1, min_row=2, max_row=current_row-1))
series9.smooth = True 
series9.graphicalProperties.line.solidFill = "008000" # Зеленый
series9.graphicalProperties.line.width = 25000 

chart9.series.append(series9)
ws9.add_chart(chart9, "D2")

# ==========================================
# ЗАДАНИЕ 10: y = 1 + 4x^2 - 2x^4/3 - КРАСНЫЙ
# ==========================================
ws10 = cast(Worksheet, wb.create_sheet(title="Задание 10"))
ws10['A1'], ws10['B1'] = 'X', 'Y'
ws10['A1'].font = ws10['B1'].font = Font(bold=True)

current_x, current_row = -3.0, 2
while current_x <= 3.01:
    ws10.cell(row=current_row, column=1, value=current_x).number_format = '0.00'
    ws10.cell(row=current_row, column=2, value=f'=1+4*A{current_row}^2-2*A{current_row}^4/3').number_format = '0.00'
    current_x = round(current_x + 0.2, 2)
    current_row += 1

chart10 = ScatterChart()
chart10.title = "y = 1 + 4x² - 2x⁴/3"
chart10.legend = None  # ОТКЛЮЧАЕМ ЛЕГЕНДУ

chart10.x_axis.scaling.min, chart10.x_axis.scaling.max = -4.0, 4.0
chart10.y_axis.scaling.min, chart10.y_axis.scaling.max = -15.0, 10.0

series10 = Series(Reference(ws10, min_col=2, min_row=2, max_row=current_row-1), 
                  Reference(ws10, min_col=1, min_row=2, max_row=current_row-1))
series10.smooth = True 
series10.graphicalProperties.line.solidFill = "FF0000" # Красный
series10.graphicalProperties.line.width = 30000

chart10.series.append(series10)
ws10.add_chart(chart10, "D2")

# ==========================================
# ФИНАЛЬНОЕ СОХРАНЕНИЕ
# ==========================================
filename = 'Практическая работа 4 (Полная).xlsx'
wb.save(filename)
print(f'Магия вне Хогвартса! Файл "{filename}" собран. Все задания с 1 по 10 готовы!')
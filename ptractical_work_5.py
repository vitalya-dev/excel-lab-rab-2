import openpyxl

# Загрузка существующего файла
try:
    wb = openpyxl.load_workbook("Фильтрация данных.xlsx")
    print("Исходный файл успешно загружен!")
except FileNotFoundError:
    print("Ошибка: Файл 'Фильтрация данных.xlsx' не найден. Убедись, что он лежит в той же папке.")
    exit()

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 1 ---
if '1_1' in wb.sheetnames:
    ws_1_1 = wb['1_1']
    # Включаем кнопочки фильтра для всей таблицы
    ws_1_1.auto_filter.ref = ws_1_1.dimensions
    print("Задание 1 (лист 1_1): Кнопки фильтра добавлены.")

if '1_2' in wb.sheetnames:
    ws_1_2 = wb['1_2']
    ws_1_2.auto_filter.ref = ws_1_2.dimensions
    print("Задание 1 (лист 1_2): Кнопки фильтра добавлены.")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 1 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 2 ---
if '2' in wb.sheetnames:
    ws_2 = wb['2']
    # Включаем фильтр для листа 2
    ws_2.auto_filter.ref = ws_2.dimensions
    print("Задание 2 (лист 2): Кнопки фильтра добавлены. (В Excel выбери Числовые фильтры -> Первые 10 -> укажи 12)")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 2 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 3 ---
if '3' in wb.sheetnames:
    ws_3 = wb['3']
    # Включаем фильтр для листа 3
    ws_3.auto_filter.ref = ws_3.dimensions
    print("Задание 3 (лист 3): Кнопки фильтра добавлены. (В Excel выбери Числовые фильтры -> Больше -> 25000)")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 3 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 4 ---
if '4' in wb.sheetnames:
    ws_4 = wb['4']
    # Включаем фильтр для листа 4, чтобы появилась возможность сортировки
    ws_4.auto_filter.ref = ws_4.dimensions
    print("Задание 4 (лист 4): Кнопки фильтра добавлены. (В Excel отсортируй по столбцу 'Торговый представитель' от А до Я)")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 4 ---

# Если в начале файла еще нет этих импортов, обязательно добавь их:
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 5 ---
if '5' in wb.sheetnames:
    ws_5 = wb['5']
    
    # 1. Настраиваем стили: желтенькая заливка, жирный шрифт, выравнивание по центру и тонкие границы
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Приятный желтый цвет
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 2. Красим шапку таблицы (1-я строка, столбцы с A до E)
    for col_idx in range(1, 6): # Столбцы от 1 (A) до 5 (E)
        header_cell = ws_5.cell(row=1, column=col_idx)
        header_cell.fill = yellow_fill
        header_cell.font = bold_font
        header_cell.alignment = center_align
        header_cell.border = thin_border
    
    # 3. Проходимся по всем строкам данных (со 2-й до конца)
    for row in range(2, ws_5.max_row + 1):
        
        # Записываем формулы
        ws_5[f'D{row}'].value = f'=IF(B{row}>10, 2, 1)'
        ws_5[f'E{row}'].value = f'=D{row}*C{row}'
        
        # Денежный формат
        currency_format = '#,##0 ₽'
        ws_5[f'C{row}'].number_format = currency_format
        ws_5[f'E{row}'].number_format = currency_format
        
        # Добавляем границы (сеточку) для каждой ячейки в строке, как на скрине
        for col_idx in range(1, 6):
            ws_5.cell(row=row, column=col_idx).border = thin_border
            
    print("Задание 5 (лист 5): Формулы, форматы и желтенькая шапка с границами успешно добавлены!")
else:
    print("Внимание: Лист '5' не найден!")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 5 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 6 ---
if '6' in wb.sheetnames:
    ws_6 = wb['6']
    
    # 1. Применяем стили к шапке (A1:D1)
    for col_idx in range(1, 5):
        cell = ws_6.cell(row=1, column=col_idx)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # 2. Обработка данных
    for row in range(2, ws_6.max_row + 1):
        # Формула скидки: если цена < 7000, то 5%, иначе 7%
        ws_6[f'C{row}'].value = f'=IF(B{row}<7000, 0.05, 0.07)'
        # Формула цены со скидкой: Цена * (1 - Скидка)
        ws_6[f'D{row}'].value = f'=B{row}*(1-C{row})'
        
        # Форматирование
        ws_6[f'B{row}'].number_format = '#,##0 ₽' # Цена
        ws_6[f'C{row}'].number_format = '0%'      # Процент скидки
        ws_6[f'D{row}'].number_format = '#,##0 ₽' # Итог
        
        # Сетка для всей строки
        for col_idx in range(1, 5):
            ws_6.cell(row=row, column=col_idx).border = thin_border
            
    print("Задание 6 (лист 6) готово: скидки посчитаны, форматы настроены.")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 6 ---


# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 7 ---
if '7' in wb.sheetnames:
    ws_7 = wb['7']
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    # 1. Стили для шапки (A1:H1)
    for col_idx in range(1, 9):
        cell = ws_7.cell(row=1, column=col_idx)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # 2. Формулы и сетка
    for row in range(2, ws_7.max_row + 1):
        # Средний балл по 5 контрольным (B-F)
        ws_7[f'G{row}'].value = f'=AVERAGE(B{row}:F{row})'
        ws_7[f'G{row}'].number_format = '0.00'
        
        # Зачет/Незачет: если средний > 3, то зачет
        ws_7[f'H{row}'].value = f'=IF(G{row}>3, "зачет", "незачет")'
        
        for col_idx in range(1, 9):
            ws_7.cell(row=row, column=col_idx).border = thin_border

    # 3. Условное форматирование для столбца H (Зачет)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Если текст равен "незачет" -> красный
    ws_7.conditional_formatting.add('H2:H100', 
        CellIsRule(operator='equal', formula=['"незачет"'], fill=red_fill))
    # Если текст равен "зачет" -> зеленый
    ws_7.conditional_formatting.add('H2:H100', 
        CellIsRule(operator='equal', formula=['"зачет"'], fill=green_fill))

    # 4. Включаем автофильтр (чтобы ты мог нажать "Сортировка А-Я" по фамилии)
    ws_7.auto_filter.ref = ws_7.dimensions

    print("Задание 7 (лист 7) готово: средний балл, зачеты и авто-покраска настроены.")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 7 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 8 ---
if '8' in wb.sheetnames:
    ws_8 = wb['8']
    from openpyxl.styles import PatternFill, Font, Border, Side
    
    # 1. Настраиваем "зеленые" стили по образцу
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid") # Тёмно-зелёный для шапки
    data_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Светло-зелёный для строк
    white_bold_font = Font(color="FFFFFF", bold=True) # Белый жирный шрифт для шапки
    black_font = Font(color="000000") # Обычный черный шрифт для данных
    
    # Тонкие границы (сеточка)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 2. Оформляем шапку (1-я строка, столбцы A - E)
    for col_idx in range(1, 6):
        cell = ws_8.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = white_bold_font
        cell.border = thin_border

    # 3. Обрабатываем данные и красим строки (со 2-й строки до конца)
    for row in range(2, ws_8.max_row + 1):
        
        # Формула для столбца D (Выполнение плана):
        # Если продажи (C) > 1 000 000, то "выполнен", иначе "не выполнен"
        ws_8[f'D{row}'].value = f'=IF(C{row}>1000000, "выполнен", "не выполнен")'
        
        # Формула для столбца E (Зарплата):
        # Если план выполнен (D), то 20000 + 5% от продаж (C), иначе просто 20000
        ws_8[f'E{row}'].value = f'=IF(D{row}="выполнен", 20000+0.05*C{row}, 20000)'
        
        # Настраиваем денежный формат (рубли без копеек) для Продаж (C) и Зарплаты (E)
        currency_format = '#,##0 ₽'
        ws_8[f'C{row}'].number_format = currency_format
        ws_8[f'E{row}'].number_format = currency_format
        
        # Применяем светло-зеленую заливку и рамки для каждой ячейки в строке данных
        for col_idx in range(1, 6):
            cell = ws_8.cell(row=row, column=col_idx)
            cell.fill = data_fill
            cell.font = black_font
            cell.border = thin_border
            
    print("Задание 8 (лист 8) готово: выполнение плана посчитано, стили 'под зеленушку' применены.")
else:
    print("Внимание: Лист '8' не найден!")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 8 ---

# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 9 ---
if '9' in wb.sheetnames:
    ws_9 = wb['9']
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.formatting.rule import CellIsRule
    
    # 1. Настраиваем стили (сделаем светло-синюю шапку для разнообразия)
    header_fill_9 = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 2. Оформляем шапку (1-я строка, столбцы A - D)
    for col_idx in range(1, 5):
        cell = ws_9.cell(row=1, column=col_idx)
        cell.fill = header_fill_9
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Данные и формулы со 2-й строки
    for row in range(2, ws_9.max_row + 1):
        # Формула И (AND): Если Рост (B) > 180 И Возраст (C) > 18
        ws_9[f'D{row}'].value = f'=IF(AND(B{row}>180, C{row}>18), "принят", "не принят")'
        
        # Рамки и выравнивание по центру для всех ячеек, кроме Фамилии
        for col_idx in range(1, 5):
            cell = ws_9.cell(row=row, column=col_idx)
            cell.border = thin_border
            if col_idx > 1: 
                cell.alignment = center_align

    # 4. Условное форматирование для наглядности (Зеленый/Красный)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws_9.conditional_formatting.add('D2:D100', CellIsRule(operator='equal', formula=['"не принят"'], fill=red_fill))
    ws_9.conditional_formatting.add('D2:D100', CellIsRule(operator='equal', formula=['"принят"'], fill=green_fill))
    
    print("Задание 9 (лист 9) готово: условия И (AND) и стили настроены.")
else:
    print("Внимание: Лист '9' не найден!")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 9 ---


# --- НАЧАЛО КОДА ДЛЯ ЗАДАНИЯ 10 ---
if '10' in wb.sheetnames:
    ws_10 = wb['10']
    
    # 1. Настраиваем стили (сделаем светло-оранжевую шапку)
    header_fill_10 = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # 2. Оформляем шапку (1-я строка, столбцы A - F)
    for col_idx in range(1, 7):
        cell = ws_10.cell(row=1, column=col_idx)
        cell.fill = header_fill_10
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Данные и формулы со 2-й строки
    for row in range(2, ws_10.max_row + 1):
        # Формула Суммы (столбец E): складываем Математику (B), Русский (C) и Ин.яз (D)
        ws_10[f'E{row}'].value = f'=SUM(B{row}:D{row})'
        
        # Формула ИЛИ (OR): Если Сумма (E) >= 180 ИЛИ Русский (C) > 80 ИЛИ Ин.яз (D) > 80
        ws_10[f'F{row}'].value = f'=IF(OR(E{row}>=180, C{row}>80, D{row}>80), "зачислен", "не зачислен")'
        
        # Рамки и выравнивание
        for col_idx in range(1, 7):
            cell = ws_10.cell(row=row, column=col_idx)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_align

    # 4. Условное форматирование для "зачислен" / "не зачислен"
    ws_10.conditional_formatting.add('F2:F100', CellIsRule(operator='equal', formula=['"не зачислен"'], fill=red_fill))
    ws_10.conditional_formatting.add('F2:F100', CellIsRule(operator='equal', formula=['"зачислен"'], fill=green_fill))

    print("Задание 10 (лист 10) готово: условия ИЛИ (OR) и стили настроены.")
else:
    print("Внимание: Лист '10' не найден!")
# --- КОНЕЦ КОДА ДЛЯ ЗАДАНИЯ 10 ---

# --- НАЧАЛО КОДА ДЛЯ ДЗ 10 ЗАДАНИЕ 1 ---
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 1. Подготавливаем данные
dz1_data = [
    ["Факс", "Персональный", 2604, 200],
    ["Факс", "Персональный+", 3774, 120],
    ["Факс", "Деловой", 2580, 160],
    ["Факс", "Профессиональный+", 4440, 400],
    ["Факс", "Профессиональный", 4500, 300],
    ["Факс", "Деловой", 1350, 230],
    ["Факс", "Профессиональный+", 6336, 190],
    ["Факс", "Профессиональный", 4920, 320],
    ["Факс", "Персональный", 2592, 543],
    ["Ксерокс", "Профессиональный+", 6168, 190],
    ["Ксерокс", "Профессиональный", 4944, 183],
    ["Ксерокс", "Профессиональный", 5520, 500],
    ["Ксерокс", "Персональный", 3780, 320],
    ["Ксерокс", "Персональный", 3828, 170],
    ["Ксерокс", "Персональный+", 6156, 400],
    ["Ксерокс", "Персональный+", 6204, 350],
    ["Ксерокс", "Деловой", 3792, 234],
    ["Ксерокс", "Деловой", 3600, 432]
]

# 2. Настраиваем стили точно по твоему скриншоту (оливково-зеленые тона)
header_fill_dz = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid") # Оливковая шапка
data_fill_dz = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")   # Светло-оливковый фон данных
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")

# Зеленые границы ячеек, как на фото
green_side = Side(style='thin', color="76933C")
green_border = Border(left=green_side, right=green_side, top=green_side, bottom=green_side)

# 3. Создаем 5 листов (база + 4 фильтра)
sheet_names = [
    "ДЗ_10_1_Фильтр_1", 
    "ДЗ_10_1_Фильтр_2", 
    "ДЗ_10_1_Фильтр_3", 
    "ДЗ_10_1_Фильтр_4"  
]

for sheet_name in sheet_names:
    ws = wb.create_sheet(title=sheet_name)
    
    # Заполняем и красим шапку
    headers = ["Товар", "Название", "Цена", "Кол-во", "Сумма"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill_dz
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = green_border
        
        # Настраиваем ширину столбцов
        if col_idx == 2:
            ws.column_dimensions[get_column_letter(col_idx)].width = 22 # Пошире для названий
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # Вставляем данные и применяем стили
    for row_idx, row_data in enumerate(dz1_data, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = green_border
            cell.fill = data_fill_dz
            cell.alignment = center_align
            
            # Формат рублей для Цены (с буквой "р.")
            if col_idx == 3: 
                cell.number_format = '#,##0"р."'
        
        # Считаем Сумму (Цена * Кол-во)
        sum_cell = ws.cell(row=row_idx, column=5)
        sum_cell.value = f'=C{row_idx}*D{row_idx}'
        sum_cell.number_format = '#,##0"р."'
        sum_cell.border = green_border
        sum_cell.fill = data_fill_dz
        sum_cell.alignment = center_align
    
    # Включаем стрелочки автофильтра на всю таблицу
    ws.auto_filter.ref = ws.dimensions

    # Твоя шпаргалка по фильтрам (когда откроешь готовый Excel-файл):

    # Лист "ДЗ_10_1_Фильтр_1": Столбец А (Товар) -> выбери «Ксерокс». Столбец B (Название) -> выбери «Персональный».

    # Лист "ДЗ_10_1_Фильтр_2": Столбец E (Сумма) -> Числовые фильтры -> Первые 10 -> исправь 10 на 8.

    # Лист "ДЗ_10_1_Фильтр_3": Столбец E (Сумма) -> Числовые фильтры -> Меньше -> впиши 1000000.

    # Лист "ДЗ_10_1_Фильтр_4": Столбец B (Название) -> оставь галочки только на «Профессиональный» и «Профессиональный+».

print("ДЗ 10 Задание 1 готово: созданы листы с таблицами и включены фильтры.")
# --- КОНЕЦ КОДА ДЛЯ ДЗ 10 ЗАДАНИЕ 1 ---

# --- НАЧАЛО КОДА ДЛЯ ДЗ 10 ЗАДАНИЕ 2 ---
# Создаем новый лист для последнего задания
ws_dz2 = wb.create_sheet(title="ДЗ_10_2_Брак")
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# 1. Настраиваем стили
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), 
    top=Side(style='thin'), bottom=Side(style='thin')
)
# Цвет шапки - песочно-желтый, как на картинке
header_fill = PatternFill(start_color="F7CB73", end_color="F7CB73", fill_type="solid")
# Ярко-желтый для плашки "Выполнил"
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

# 2. Оформляем мини-таблицу слева сверху (строки 1 и 2)
ws_dz2['A1'] = "Цех"
ws_dz2['B1'] = "№3"
ws_dz2['A2'] = "Дата"
ws_dz2['B2'] = "17.03.2016" # Оставим дату как на скрине

for r in range(1, 3):
    for c in range(1, 3):
        ws_dz2.cell(row=r, column=c).border = thin_border

# 3. Желтая плашка "Выполнил..."
ws_dz2.merge_cells('D2:F2') # Объединяем, чтобы влезло красиво
ws_dz2['D2'] = "Выполнил Козлов А.В."
ws_dz2['D2'].fill = yellow_fill
ws_dz2['D2'].alignment = center_align

# 4. Заголовки основной таблицы (строка 4)
headers = [
    "Название\nдетали", "Кол-во,\nшт", "Брак, шт", "Себесто\nимость", 
    "Сумма", "Брак, %", "Штраф", "Сумма за\nвычетом\nштрафа"
]

for col_idx, h in enumerate(headers, start=1):
    cell = ws_dz2.cell(row=4, column=col_idx, value=h)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    
# Настраиваем ширину столбцов, чтобы текст переносился как надо
ws_dz2.column_dimensions['A'].width = 12
ws_dz2.column_dimensions['B'].width = 10
ws_dz2.column_dimensions['C'].width = 10
ws_dz2.column_dimensions['D'].width = 12
ws_dz2.column_dimensions['E'].width = 10
ws_dz2.column_dimensions['F'].width = 10
ws_dz2.column_dimensions['G'].width = 10
ws_dz2.column_dimensions['H'].width = 14

# 5. Исходные данные
data = [
    ["Шайба", 120, 20, 10],
    ["Винт", 100, 25, 14],
    ["Гайка", 115, 10, 16],
    ["Болт", 95, 27, 10],
    ["Шуруп", 87, 12, 15]
]

# 6. Заполняем данные и прописываем формулы
start_row = 5
for i, row_data in enumerate(data):
    current_row = start_row + i
    
    # Вставляем текстовые и числовые значения
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws_dz2.cell(row=current_row, column=col_idx, value=val)
        cell.border = thin_border
        if col_idx > 1:
            cell.alignment = center_align

    # Формула: Сумма (E) = Кол-во (B) * Себестоимость (D)
    ws_dz2[f'E{current_row}'] = f'=B{current_row}*D{current_row}'
    ws_dz2[f'E{current_row}'].border = thin_border
    ws_dz2[f'E{current_row}'].alignment = center_align
    
    # Формула: Брак % (F) = Брак шт (C) / Кол-во (B)
    ws_dz2[f'F{current_row}'] = f'=C{current_row}/B{current_row}'
    ws_dz2[f'F{current_row}'].number_format = '0%' # Превращаем десятичное число в проценты
    ws_dz2[f'F{current_row}'].border = thin_border
    ws_dz2[f'F{current_row}'].alignment = center_align
    
    # Формула: Штраф (G). Важно: в Excel 10% - это 0.1
    # Если процент брака (F) > 10% (0.1), то 5% (0.05) * Сумма (E), иначе 0
    ws_dz2[f'G{current_row}'] = f'=IF(F{current_row}>0.1, 0.05*E{current_row}, 0)'
    ws_dz2[f'G{current_row}'].border = thin_border
    ws_dz2[f'G{current_row}'].alignment = center_align
    
    # Формула: Сумма за вычетом (H) = Сумма (E) - Штраф (G)
    ws_dz2[f'H{current_row}'] = f'=E{current_row}-G{current_row}'
    ws_dz2[f'H{current_row}'].border = thin_border
    ws_dz2[f'H{current_row}'].alignment = center_align

# 7. Финальные строки "Итого" и "К выдаче"
total_row = start_row + len(data)      # 10-я строка
payout_row = start_row + len(data) + 1  # 11-я строка

# Объединяем ячейки A-G для "Итого"
ws_dz2.merge_cells(f'A{total_row}:G{total_row}')
ws_dz2[f'A{total_row}'] = "Итого"
ws_dz2[f'A{total_row}'].alignment = right_align
# Сумма по столбцу H
ws_dz2[f'H{total_row}'] = f'=SUM(H{start_row}:H{total_row-1})'

# Объединяем ячейки A-G для "К выдаче"
ws_dz2.merge_cells(f'A{payout_row}:G{payout_row}')
ws_dz2[f'A{payout_row}'] = "К выдаче"
ws_dz2[f'A{payout_row}'].alignment = right_align
# 7% от Итого (H10)
ws_dz2[f'H{payout_row}'] = f'=0.07*H{total_row}'

# Рисуем рамки для объединенных нижних строк (openpyxl требует обводить каждую ячейку в слиянии)
for r in [total_row, payout_row]:
    for c in range(1, 9):
        ws_dz2.cell(row=r, column=c).border = thin_border
        if c == 8: # Выравниваем результаты
            ws_dz2.cell(row=r, column=c).alignment = center_align

print("ДЗ 10 Задание 2 готово: расчет брака, штрафов и итогов завершен!")
# --- КОНЕЦ КОДА ДЛЯ ДЗ 10 ЗАДАНИЕ 2 ---

# Сохранение финального результата
output_filename = 'Практическая_работа_5.xlsx'
wb.save(output_filename)
print(f'\nВсе готово, бро! Файл "{output_filename}" успешно сохранен.')